"""Generate sample_data/messy_sales_example.xlsx.

The workbook is deliberately "messy" so it exercises every acceptance criterion:

* a title row above the header row in the data sheet (header NOT at row 0)
* a measure column whose name won't exact-match the natural guideline term
  (Total_Sales_USD vs "revenue")
* a categorical column (Product) with 50+ unique values to trigger bucketing
* an "Instructions" sheet with a line referencing a column that does not exist
  ("profit margin") to test the skip-and-flag path

Run:  python sample_data/make_sample_data.py
"""

from __future__ import annotations

import random
from pathlib import Path

from openpyxl import Workbook

PRODUCTS = [
    f"Widget {name}"
    for name in (
        "Alpha", "Beta", "Gamma", "Delta", "Epsilon", "Zeta", "Eta", "Theta", "Iota", "Kappa",
        "Lambda", "Mu", "Nu", "Xi", "Omicron", "Pi", "Rho", "Sigma", "Tau", "Upsilon",
        "Phi", "Chi", "Psi", "Omega", "Aurora", "Breeze", "Comet", "Drift", "Ember", "Frost",
        "Gale", "Halo", "Iris", "Juno", "Kite", "Lumen", "Mist", "North", "Oasis", "Pulse",
        "Quill", "Rune", "Storm", "Tide", "Umbra", "Vale", "Wisp", "Yard", "Zenith", "Glint",
        "Nova", "Orbit", "Prism", "Quest", "Rider", "Sable", "Trail", "Urban", "Vortex", "Willow",
    )
]

REGIONS = ["North", "South", "East", "West"]
MONTHS = [f"{y}-{m:02d}-01" for y in (2024, 2025) for m in range(1, 13)]

INSTRUCTIONS = [
    "Plot total revenue over time (line chart)",
    "Show sales by region as a bar chart",
    "Pie chart of units sold by product",
    "Bar chart of profit margin by product",
    "Scatter plot of units sold vs total sales",
    "Horizontal bar of top products by revenue",
]


def build() -> None:
    random.seed(42)
    wb = Workbook()

    ws = wb.active
    ws.title = "Sales"
    ws.append(["Quarterly Sales Report"])  # row 0: title, NOT the header
    ws.append(["Date", "Region", "Product", "Total_Sales_USD", "Units_Sold", "Cost_USD"])

    for _ in range(400):
        date = MONTHS[random.randrange(len(MONTHS))]
        region = random.choice(REGIONS)
        product = random.choice(PRODUCTS)
        unit_price = random.uniform(8.0, 120.0)
        units = random.randint(1, 60)
        row = [
            date[:10],
            region,
            product,
            round(units * unit_price, 2),  # Total_Sales_USD  <-> "revenue" in the guideline
            units,
            round(units * unit_price * random.uniform(0.45, 0.7), 2),
        ]
        ws.append(row)

    inst = wb.create_sheet("Instructions")
    for i, line in enumerate(INSTRUCTIONS, start=1):
        inst.cell(row=i, column=1, value=line)

    out = Path(__file__).resolve().parent / "messy_sales_example.xlsx"
    wb.save(out)
    print("written:", out)


if __name__ == "__main__":
    build()