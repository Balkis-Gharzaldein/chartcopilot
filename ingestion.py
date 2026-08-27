"""Excel ingestion + per-sheet header detection + schema profiling.

Only SheetProfile objects (column names, dtypes, small sample values, null and
unique counts) ever leave this module -- full DataFrames stay in memory for the
agent's tool-calling loop and are never sent to an LLM.
"""

from __future__ import annotations

import io
from dataclasses import dataclass

import openpyxl
import pandas as pd

from schemas import ColumnProfile, SheetProfile

HEADER_SCAN_ROWS = 10
SAMPLE_SIZE = 5


@dataclass
class Workbook:
    profiles: list[SheetProfile]
    frames: dict[str, pd.DataFrame]  # sheet_name -> usable dataframe
    raw: "openpyxl.Workbook | None" = None  # openpyxl workbook, for guideline cells

    def profile_for(self, sheet: str) -> SheetProfile:
        for p in self.profiles:
            if p.sheet_name == sheet:
                return p
        raise KeyError(f"No sheet named {sheet!r} in workbook")


def detect_header_row(rows: list[list], max_scan: int = HEADER_SCAN_ROWS) -> int:
    """Pick the row (index) most likely to hold column headers.

    Heuristic: score each of the first `max_scan` rows by how "header-like" it
    is -- high ratio of non-null string cells, short single-word names (typical
    of column headers rather than prose), and a good number of distinct values.
    Falls back to 0 if nothing useful is found.
    """
    best_idx, best_score = 0, -1.0
    for i in range(min(len(rows), max_scan)):
        vals = rows[i] or []
        non_null = [v for v in vals if v is not None and str(v).strip() != ""]
        if not non_null:
            continue
        str_ratio = sum(1 for v in non_null if isinstance(v, str)) / len(non_null)
        header_like = sum(
            1
            for v in non_null
            if isinstance(v, str) and " " not in str(v).strip()
        ) / len(non_null)
        distinct = len({str(v).strip().lower() for v in non_null}) / len(non_null)
        score = str_ratio + 0.4 * header_like + 0.2 * distinct
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx


def _clean_headers(headers: list) -> list[str]:
    cleaned = []
    seen = set()
    for idx, h in enumerate(headers):
        if h is None or str(h).strip() == "":
            name = f"unnamed_{idx}"
        else:
            name = str(h).strip()
        name = name.replace("\n", " ").replace("\r", " ")
        if name in seen:
            name = f"{name}_{idx}"
        seen.add(name)
        cleaned.append(name)
    return cleaned


def profile_sheet(sheet_name: str, df: pd.DataFrame, raw_rows: list[list]) -> SheetProfile:
    cols: list[ColumnProfile] = []
    for name in df.columns:
        series = df[name]
        sample = [
            str(v) if v is not None else ""
            for v in series.dropna().astype(str).head(SAMPLE_SIZE).tolist()
        ]
        cols.append(
            ColumnProfile(
                name=str(name),
                dtype=str(series.dtype),
                sample_values=sample,
                null_count=int(series.isna().sum()),
                unique_count=int(series.nunique(dropna=True)),
            )
        )
    return SheetProfile(sheet_name=sheet_name, columns=cols, row_count=int(len(df)))


def _load_sheets(file_bytes: bytes, source_name: str = "upload") -> Workbook:
    """Parse every sheet, detect its header row, and build a clean DataFrame + profile."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    sheet_names = wb.sheetnames

    profiles: list[SheetProfile] = []
    frames: dict[str, pd.DataFrame] = {}

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        raw = [list(row) for row in ws.iter_rows(values_only=True)]
        header_idx = detect_header_row(raw)
        header = raw[header_idx] if raw else []
        clean = _clean_headers(header)
        data_rows = raw[header_idx + 1 :]
        data_rows = [r for r in data_rows if any(v is not None for v in r)]
        if not clean or not data_rows:
            # Empty or header-only sheet: emit an empty frame + profile.
            df = pd.DataFrame()
            profiles.append(SheetProfile(sheet_name=sheet_name, columns=[], row_count=0))
            frames[sheet_name] = df
            continue

        if len(clean) > len(data_rows[0]):
            clean = clean[: len(data_rows[0])]
        df = pd.DataFrame(data_rows, columns=clean[: len(data_rows[0])])
        profiles.append(profile_sheet(sheet_name, df, raw))
        frames[sheet_name] = df

    return Workbook(profiles=profiles, frames=frames, raw=wb)


def ingest_file(file_bytes: bytes, source_name: str = "upload") -> Workbook:
    if source_name and source_name.lower().endswith(".csv"):
        return _ingest_csv(file_bytes, source_name)
    return _load_sheets(file_bytes, source_name)


def _ingest_csv(file_bytes: bytes, source_name: str) -> Workbook:
    df = pd.read_csv(io.BytesIO(file_bytes))
    df.columns = _clean_headers(list(df.columns))
    profile = profile_sheet("data", df, None)
    return Workbook(profiles=[profile], frames={"data": df}, raw=None)